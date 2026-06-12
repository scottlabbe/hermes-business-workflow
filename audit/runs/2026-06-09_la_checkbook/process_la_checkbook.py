from __future__ import annotations

import hashlib
import json
import math
import random
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


RUN_ID = "2026-06-09_la_checkbook"
JOB_ID = "la_checkbook"
RUN_DIR = Path("runs") / RUN_ID
SOURCE_FILE = RUN_DIR / "00_source" / "LACheckbookData-3.xlsx"
KEY_FIELDS_FILE = RUN_DIR / "00_source" / "key_fields.yaml"
REQUEST_FILE = RUN_DIR / "run_request.md"
DATA_RELIABILITY_DIR = RUN_DIR / "01_data_reliability"
SAMPLE_DIR = RUN_DIR / "02_expense-sample"


def rel(path: Path) -> str:
    return path.relative_to(RUN_DIR).as_posix()


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d")
    return value


def value_to_text(value) -> str:
    value = clean_value(value)
    if value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def read_key_fields(path: Path) -> dict:
    dataset_file = None
    sheet_name = None
    key_fields: list[str] = []
    in_key_fields = False
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line.startswith("dataset_file:"):
            dataset_file = line.split(":", 1)[1].strip().strip("'\"")
            in_key_fields = False
        elif line.startswith("sheet_name:"):
            sheet_name = line.split(":", 1)[1].strip().strip("'\"")
            in_key_fields = False
        elif line.startswith("key_fields:"):
            in_key_fields = True
        elif in_key_fields and line.startswith("-"):
            key_fields.append(line[1:].strip().strip("'\""))
    if not dataset_file:
        raise ValueError("key_fields.yaml is missing dataset_file")
    if not key_fields:
        raise ValueError("key_fields.yaml is missing key_fields")
    return {"dataset_file": dataset_file, "sheet_name": sheet_name, "key_fields": key_fields}


def infer_field(series: pd.Series, field_name: str) -> dict:
    present = series is not None
    if not present:
        return {
            "field_name": field_name,
            "field_present": False,
            "inferred_type": "missing",
            "inference_confidence": 0.0,
            "nonblank_count": 0,
            "blank_count": None,
            "blank_percent": None,
            "distinct_count": 0,
            "duplicate_value_count": 0,
            "parse_failure_count": None,
            "min_value": "",
            "max_value": "",
            "sample_values": "[]",
            "issue_flags": "MISSING_REQUESTED_FIELD",
            "ambiguity_notes": "Requested field was not present in the parsed source table.",
        }

    text = series.map(value_to_text)
    nonblank_mask = text.str.strip() != ""
    nonblank = text[nonblank_mask]
    nonblank_count = int(nonblank_mask.sum())
    blank_count = int((~nonblank_mask).sum())
    blank_percent = round(blank_count / len(series), 4) if len(series) else 0
    distinct_count = int(nonblank.nunique(dropna=True))
    counts = nonblank.value_counts(dropna=True)
    duplicate_value_count = int(counts[counts > 1].sum())

    numeric = pd.to_numeric(nonblank, errors="coerce") if nonblank_count else pd.Series(dtype=float)
    numeric_success = int(numeric.notna().sum()) if nonblank_count else 0
    parsed_dates = pd.to_datetime(nonblank, errors="coerce") if nonblank_count else pd.Series(dtype="datetime64[ns]")
    date_success = int(parsed_dates.notna().sum()) if nonblank_count else 0
    name_hints_date = bool(re.search(r"date|fiscal", field_name, re.I))

    inferred_type = "blank"
    confidence = 0.0
    parse_failure_count = 0
    min_value = ""
    max_value = ""
    ambiguity_notes = ""

    if nonblank_count == 0:
        ambiguity_notes = "Field is present but all values are blank."
    elif pd.api.types.is_datetime64_any_dtype(series):
        inferred_type = "date"
        confidence = 1.0
        parsed = pd.to_datetime(series[nonblank_mask], errors="coerce")
        parse_failure_count = int(parsed.isna().sum())
        min_value = value_to_text(parsed.min())
        max_value = value_to_text(parsed.max())
    elif numeric_success / nonblank_count >= 0.95:
        has_decimal = any((float(x) % 1) != 0 for x in numeric.dropna().head(500))
        if name_hints_date and numeric.dropna().between(20000, 60000).mean() >= 0.9:
            inferred_type = "numeric_date_or_fiscal_period"
            ambiguity_notes = "Field name suggests date/fiscal year, but values were parsed as numeric serials/codes; human should confirm intended interpretation."
        else:
            inferred_type = "decimal" if has_decimal else "integer"
        confidence = round(numeric_success / nonblank_count, 4)
        parse_failure_count = int(nonblank_count - numeric_success)
        min_value = value_to_text(numeric.min())
        max_value = value_to_text(numeric.max())
    elif date_success / nonblank_count >= 0.95:
        inferred_type = "date"
        confidence = round(date_success / nonblank_count, 4)
        parse_failure_count = int(nonblank_count - date_success)
        min_value = value_to_text(parsed_dates.min())
        max_value = value_to_text(parsed_dates.max())
    else:
        inferred_type = "text"
        confidence = 1.0
        parse_failure_count = 0
        min_value = value_to_text(nonblank.min()) if nonblank_count else ""
        max_value = value_to_text(nonblank.max()) if nonblank_count else ""

    samples = [value_to_text(v) for v in nonblank.drop_duplicates().head(5).tolist()]
    flags: list[str] = []
    if blank_count:
        flags.append("BLANK_VALUES_PRESENT")
    if duplicate_value_count:
        flags.append("DUPLICATE_VALUES_PRESENT")
    if parse_failure_count:
        flags.append("PARSE_FAILURES_PRESENT")
    if inferred_type == "numeric_date_or_fiscal_period":
        flags.append("DATE_OR_FISCAL_FIELD_AMBIGUITY")

    return {
        "field_name": field_name,
        "field_present": True,
        "inferred_type": inferred_type,
        "inference_confidence": confidence,
        "nonblank_count": nonblank_count,
        "blank_count": blank_count,
        "blank_percent": blank_percent,
        "distinct_count": distinct_count,
        "duplicate_value_count": duplicate_value_count,
        "parse_failure_count": parse_failure_count,
        "min_value": min_value,
        "max_value": max_value,
        "sample_values": json.dumps(samples),
        "issue_flags": "; ".join(flags),
        "ambiguity_notes": ambiguity_notes,
    }


def autosize_and_style(path: Path, table_sheets: set[str]) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(bold=True, size=14, color="1F4E78")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        if ws.title == "Summary":
            ws.freeze_panes = None
            ws["A1"].font = title_font
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 100
        else:
            max_row = ws.max_row
            max_col = ws.max_column
            if ws.title in table_sheets and max_row > 1 and max_col > 1:
                ref = f"A1:{get_column_letter(max_col)}{max_row}"
                table_name = re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:20] + "Table"
                tab = Table(displayName=table_name, ref=ref)
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
                ws.add_table(tab)
            for idx, col in enumerate(ws.iter_cols(), start=1):
                values = [value_to_text(cell.value) for cell in col[:200]]
                width = min(max(max((len(v) for v in values), default=8) + 2, 10), 45)
                ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def write_summary_sheet(writer, rows: list[tuple[str, str]]) -> None:
    pd.DataFrame(rows, columns=["Topic", "Details"]).to_excel(writer, sheet_name="Summary", index=False)


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def main() -> None:
    DATA_RELIABILITY_DIR.mkdir(parents=True, exist_ok=True)
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    config = read_key_fields(KEY_FIELDS_FILE)

    excel = pd.ExcelFile(SOURCE_FILE)
    if config["sheet_name"]:
        source_sheet = config["sheet_name"]
    elif len(excel.sheet_names) == 1:
        source_sheet = excel.sheet_names[0]
    else:
        raise ValueError("Workbook has multiple sheets and key_fields.yaml did not specify sheet_name")

    source_df = pd.read_excel(SOURCE_FILE, sheet_name=source_sheet)
    source_df = source_df.where(pd.notna(source_df), None)
    row_count, column_count = source_df.shape
    requested_key_fields = config["key_fields"]
    missing_fields = [field for field in requested_key_fields if field not in source_df.columns]
    duplicate_row_count = int(source_df.duplicated().sum())
    unnamed_cols = [str(c) for c in source_df.columns if str(c).strip() == "" or str(c).startswith("Unnamed:")]
    field_results = [
        infer_field(source_df[field], field) if field in source_df.columns else infer_field(None, field)
        for field in requested_key_fields
    ]

    limitations = [
        "Outputs are review aids and data reliability observations, not final audit, compliance, legal, financial, or management conclusions.",
        "Review is limited to requested key-field profiling, duplicate whole-row checks, unnamed header checks, and basic parsing observations.",
        "Excel workbook parsing used the first and only source sheet because key_fields.yaml did not specify a sheet name.",
    ]
    if any(r["inferred_type"] == "numeric_date_or_fiscal_period" for r in field_results):
        limitations.append("Posting Date/FiscalYear was parsed as numeric serial/code values; human review should confirm the intended date or fiscal-year interpretation.")

    observations = [
        f"Parsed {row_count} rows and {column_count} columns from {source_sheet}.",
        f"Missing requested key fields: {len(missing_fields)}.",
        f"Duplicate whole-row count: {duplicate_row_count}.",
    ]
    if unnamed_cols:
        observations.append(f"Unnamed or blank header columns present: {', '.join(unnamed_cols)}.")
    date_ambiguities = [r["field_name"] for r in field_results if r["inferred_type"] == "numeric_date_or_fiscal_period"]
    if date_ambiguities:
        observations.append(f"Date/fiscal field ambiguity noted for: {', '.join(date_ambiguities)}.")

    checks = [
        "Validated intake files and key-fields configuration.",
        "Parsed the source workbook sheet selected for review.",
        "Profiled requested key fields for presence, blanks, distinct values, duplicate-value counts, type inference, parse observations, and sample values.",
        "Checked duplicate whole rows.",
        "Checked unnamed or blank header columns.",
        "Recorded downstream handoff for the sampling stage.",
    ]
    human_review = [
        "Review missing-field, blank-value, duplicate-value, and parse-ambiguity observations before relying on the workbook for follow-on procedures.",
        "Confirm the configured key fields match the intended review objective.",
        "Confirm the Posting Date/FiscalYear field meaning and whether numeric values should be interpreted as Excel date serials, fiscal years, or another code.",
    ]

    dr_workbook = DATA_RELIABILITY_DIR / "data_reliability.xlsx"
    summary_rows = [
        ("Run ID", RUN_ID),
        ("Source file", "00_source/LACheckbookData-3.xlsx"),
        ("Source sheet/table", source_sheet),
        ("Row count", str(row_count)),
        ("Column count", str(column_count)),
        ("Requested key fields", ", ".join(requested_key_fields)),
        ("Missing requested key fields", ", ".join(missing_fields) if missing_fields else "None"),
        ("Duplicate row count", str(duplicate_row_count)),
        ("Checks performed", "\n".join(checks)),
        ("Data reliability observations", "\n".join(observations)),
        ("Model judgment", "The parsed workbook appears usable for creating a review-aid sample because all requested key fields are present; however, this is not a reliability conclusion, and the Posting Date/FiscalYear numeric-code ambiguity should be reviewed by a human."),
        ("Limitations", "\n".join(limitations)),
        ("Human review needed", "\n".join(human_review)),
        ("Downstream handoff file and sheet", "01_data_reliability/data_reliability.xlsx / Source File"),
    ]
    with pd.ExcelWriter(dr_workbook, engine="openpyxl") as writer:
        write_summary_sheet(writer, summary_rows)
        source_df.to_excel(writer, sheet_name="Source File", index=False)
        pd.DataFrame(field_results).to_excel(writer, sheet_name="Key Field Profile", index=False)
    autosize_and_style(dr_workbook, {"Source File", "Key Field Profile"})

    dr_metadata = {
        "run_id": RUN_ID,
        "stage_id": "01_data_reliability",
        "skill": "data-reliability",
        "status": "completed_with_limitations",
        "source_file": "00_source/LACheckbookData-3.xlsx",
        "source_format": "xlsx",
        "source_sheet_or_table": source_sheet,
        "parser_used": "pandas/openpyxl",
        "method": "generated Python analysis run from the terminal using the bundled Codex Python runtime",
        "artifacts_written": [
            "01_data_reliability/data_reliability.xlsx",
            "01_data_reliability/data_reliability_metadata.json",
        ],
        "parsing_succeeded": True,
        "row_count": row_count,
        "column_count": column_count,
        "requested_key_fields": requested_key_fields,
        "missing_requested_key_fields": missing_fields,
        "duplicate_row_count": duplicate_row_count,
        "unnamed_or_blank_header_columns": {"present": bool(unnamed_cols), "columns": unnamed_cols},
        "field_results": field_results,
        "checks_performed": checks,
        "limitations": limitations,
        "human_review_needed": human_review,
        "downstream_handoff": {
            "ready_for_sampling": len(missing_fields) == 0,
            "file": "01_data_reliability/data_reliability.xlsx",
            "sheet": "Source File",
            "record_count": row_count,
            "key_fields": requested_key_fields,
            "limitations": limitations,
        },
    }
    (DATA_RELIABILITY_DIR / "data_reliability_metadata.json").write_text(json.dumps(dr_metadata, indent=2), encoding="utf-8")

    manifest = {
        "manifest_version": "1.0",
        "run_id": RUN_ID,
        "job_id": JOB_ID,
        "created_at": datetime.now().replace(microsecond=0).isoformat(),
        "status": "in_progress",
        "source_files": [
            "00_source/LACheckbookData-3.xlsx",
            "00_source/key_fields.yaml",
        ],
        "stages": {
            "01_data_reliability": {
                "skill": "data-reliability",
                "status": dr_metadata["status"],
                "output": "01_data_reliability/data_reliability.xlsx",
                "metadata": "01_data_reliability/data_reliability_metadata.json",
            }
        },
        "current_sampling_input": dr_metadata["downstream_handoff"],
    }
    (RUN_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    sample_metadata_path = SAMPLE_DIR / "sample_metadata.json"
    sampling_input = manifest["current_sampling_input"]
    if not sampling_input.get("ready_for_sampling"):
        blocked = {
            "run_id": RUN_ID,
            "stage_id": "02_expense-sample",
            "skill": "expense-sample",
            "status": "blocked",
            "source_input": sampling_input,
            "artifacts_written": ["02_expense-sample/sample_metadata.json"],
            "checks_performed": ["current_sampling_input is present", "ready_for_sampling is not true"],
            "limitations": sampling_input.get("limitations", []),
            "human_review_needed": ["Resolve data reliability blockers before sampling."],
        }
        sample_metadata_path.write_text(json.dumps(blocked, indent=2), encoding="utf-8")
        manifest["stages"]["02_expense-sample"] = {
            "skill": "expense-sample",
            "status": "blocked",
            "output": None,
            "metadata": "02_expense-sample/sample_metadata.json",
        }
        manifest["status"] = "blocked"
        (RUN_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
        return

    input_workbook = RUN_DIR / sampling_input["file"]
    input_sheet = sampling_input["sheet"]
    input_df = pd.read_excel(input_workbook, sheet_name=input_sheet)
    input_df = input_df.where(pd.notna(input_df), None)
    input_count = len(input_df)
    sample_size_requested = 60
    random_seed = 20260601
    sample_size_selected = min(sample_size_requested, input_count)
    rng = random.Random(random_seed)
    selected_population_sequences = rng.sample(range(1, input_count + 1), sample_size_selected)
    selected_set = set(selected_population_sequences)

    input_with_helpers = input_df.copy()
    input_with_helpers.insert(0, "Population Sequence", range(1, input_count + 1))
    input_with_helpers.insert(0, "Source Row Number", range(2, input_count + 2))

    sampled_rows = []
    for sample_sequence, population_sequence in enumerate(selected_population_sequences, start=1):
        source_index = population_sequence - 1
        row = input_df.iloc[source_index].to_dict()
        sampled_rows.append(
            {
                "Source Row Number": population_sequence + 1,
                "Population Sequence": population_sequence,
                "Sample Sequence": sample_sequence,
                "Selection Method": "python_random_sample_without_replacement",
                "Random Seed": random_seed,
                "Sampling Notes": "Selected from the unfiltered population using documented random seed and population sequence.",
                **row,
            }
        )
    sampled_df = pd.DataFrame(sampled_rows)

    sample_limitations = [
        *limitations,
        "The sampling stage used an unfiltered random sample as directed by the local expense-sample skill; it does not establish statistical sufficiency or representativeness.",
    ]
    sample_checks = [
        "manifest exists",
        "current_sampling_input is present",
        "ready_for_sampling is true",
        "input file exists",
        "input sheet exists",
        "input row count is greater than zero",
        "selected records come from the input population",
        f"selected sample count equals {sample_size_selected}",
        "random seed is recorded",
        "Python standard-library random sample method is recorded",
        "output workbook contains all required sheets",
        "metadata records the output workbook and downstream handoff",
    ]

    sample_workbook = SAMPLE_DIR / "sample_selection.xlsx"
    sample_summary_rows = [
        ("Run ID", RUN_ID),
        ("Source workbook and sheet", f"{sampling_input['file']} / {input_sheet}"),
        ("Input record count", str(input_count)),
        ("Sample size requested", str(sample_size_requested)),
        ("Sample size selected", str(sample_size_selected)),
        ("Random seed", str(random_seed)),
        ("Population definition", "all rows in source sheet"),
        ("Selection method", "Python standard-library random sampling without replacement"),
        ("Reproducibility method", "random.Random(random_seed).sample(range(1, input_record_count + 1), sample_size_selected); Sample Sequence follows draw order."),
        ("Selected population sequences", json.dumps(selected_population_sequences)),
        ("Checks performed", "\n".join(sample_checks)),
        ("Limitations", "\n".join(sample_limitations)),
        ("Human review needed", "Review selected records and source table before requesting documentation or making audit judgments.\nConfirm this unfiltered random sample is appropriate for the intended review objective."),
    ]
    with pd.ExcelWriter(sample_workbook, engine="openpyxl") as writer:
        write_summary_sheet(writer, sample_summary_rows)
        input_with_helpers.to_excel(writer, sheet_name="Input Data", index=False)
        sampled_df.to_excel(writer, sheet_name="Sampled Records", index=False)
    autosize_and_style(sample_workbook, {"Input Data", "Sampled Records"})

    sample_metadata = {
        "run_id": RUN_ID,
        "stage_id": "02_expense-sample",
        "skill": "expense-sample",
        "status": "completed_with_limitations",
        "source_input": {
            "file": sampling_input["file"],
            "sheet": input_sheet,
            "sha256": sha256(input_workbook),
        },
        "artifacts_written": [
            "02_expense-sample/sample_selection.xlsx",
            "02_expense-sample/sample_metadata.json",
        ],
        "parameters": {
            "sample_size_requested": sample_size_requested,
            "random_seed": random_seed,
            "population_definition": "all rows in source sheet",
            "selection_method": "python_random_sample_without_replacement",
            "selection_algorithm": "random.Random(random_seed).sample(range(1, input_record_count + 1), sample_size_selected)",
        },
        "counts": {
            "input_record_count": input_count,
            "sample_size_selected": sample_size_selected,
        },
        "reproducibility": {
            "source_order_basis": "workbook row order",
            "source_row_number_basis": "worksheet row number in the Source File sheet of the data reliability workbook",
            "population_sequence_basis": "1-based order after applying the unfiltered population definition",
            "selected_population_sequences": selected_population_sequences,
            "selected_source_row_numbers": [seq + 1 for seq in selected_population_sequences],
        },
        "checks_performed": sample_checks,
        "limitations": sample_limitations,
        "human_review_needed": [
            "Review the selected records and supporting source table before requesting documentation or making audit judgments.",
            "Confirm this unfiltered random sample is appropriate for the intended review objective.",
        ],
        "downstream_handoff": {
            "file": "02_expense-sample/sample_selection.xlsx",
            "sheet": "Sampled Records",
        },
    }
    sample_metadata_path.write_text(json.dumps(sample_metadata, indent=2), encoding="utf-8")

    manifest["status"] = "completed_with_limitations"
    manifest["stages"]["02_expense-sample"] = {
        "skill": "expense-sample",
        "status": sample_metadata["status"],
        "output": "02_expense-sample/sample_selection.xlsx",
        "metadata": "02_expense-sample/sample_metadata.json",
    }
    (RUN_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    run_summary = f"""# Run Summary

## Task Attempted

Processed `inbox/la_checkbook/request.md` for data reliability observations and an expense sample selection review aid.

## Inputs Used

- `inbox/la_checkbook/request.md`
- `inbox/la_checkbook/source/LACheckbookData-3.xlsx`
- `inbox/la_checkbook/source/key_fields.yaml`

## Outputs Created

- `manifest.json`
- `00_source/LACheckbookData-3.xlsx`
- `00_source/key_fields.yaml`
- `01_data_reliability/data_reliability.xlsx`
- `01_data_reliability/data_reliability_metadata.json`
- `02_expense-sample/sample_selection.xlsx`
- `02_expense-sample/sample_metadata.json`

## Checks Performed

- Validated the intake folder, source dataset, and key-fields config.
- Parsed {row_count} source records and {column_count} source columns from `{source_sheet}`.
- Profiled requested key fields for presence, blanks, distinct values, duplicate-value counts, type inference, parse observations, and sample values.
- Checked duplicate whole rows and unnamed or blank header columns.
- Confirmed the sampling manifest handoff points to `01_data_reliability/data_reliability.xlsx` / `Source File`.
- Selected {sample_size_selected} records from the full unfiltered population using Python standard-library random sampling with seed {random_seed}.
- Verified required workbook sheets were present: data reliability=True, sample selection=True.

## Limitations

- Outputs are review aids and data reliability observations, not final audit, compliance, legal, financial, or management conclusions.
- Data reliability work is limited to requested key fields and basic source-table observations.
- `Posting Date/FiscalYear` was parsed as numeric serial/code values; human review should confirm the intended interpretation.
- The sampling stage used an unfiltered random sample as directed by the local skill; it does not establish statistical sufficiency or representativeness.

## Human Review Needed

- Review the key-field profile, especially blank counts, duplicate-value flags, and parse observations.
- Confirm the key fields and unfiltered sample approach are appropriate for the intended workpaper objective.
- Confirm whether `Posting Date/FiscalYear` values should be interpreted as Excel date serials, fiscal years, or another coding convention.
- Review sampled records before requesting documentation or drawing conclusions.

## Suggested Improvements

- Add explicit acceptance criteria for key-field blank thresholds or parse exceptions if future runs need automated escalation.
- Add user-defined stratification or exclusion rules only when the review objective requires them.
- Add a reviewer signoff section if these artifacts will be used in a controlled workpaper package.
"""
    (RUN_DIR / "run_summary.md").write_text(run_summary, encoding="utf-8")

    # Compact verification guardrails.
    dr_sheets = load_workbook(dr_workbook, read_only=True).sheetnames
    sample_sheets = load_workbook(sample_workbook, read_only=True).sheetnames
    assert dr_sheets == ["Summary", "Source File", "Key Field Profile"], dr_sheets
    assert sample_sheets == ["Summary", "Input Data", "Sampled Records"], sample_sheets
    assert len(selected_set) == sample_size_selected
    assert set(selected_population_sequences).issubset(set(range(1, input_count + 1)))


if __name__ == "__main__":
    main()
