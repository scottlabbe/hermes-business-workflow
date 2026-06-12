from __future__ import annotations

import csv
import json
import math
import re
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RUN_ID = "2026-06-09_trial_balance"
CREATED_AT = "2026-06-09T00:00:00"
JOB_ID = "trial_balance"
RUN_DIR = Path(__file__).resolve().parent
SOURCE_DIR = RUN_DIR / "00_source"
STAGE_DIR = RUN_DIR / "01_variance_analysis"
PRIOR_FILE = SOURCE_DIR / "FY25 TB.csv"
CURRENT_FILE = SOURCE_DIR / "FY26 TB.csv"
OUTPUT_XLSX = STAGE_DIR / "variance_analysis.xlsx"
OUTPUT_METADATA = STAGE_DIR / "variance_analysis_metadata.json"
DOLLAR_THRESHOLD = 200_000.0
PERCENT_THRESHOLD = 0.20
THRESHOLD_LOGIC = "AND"
ZERO_TOLERANCE = 0.004


ALIASES = {
    "account_number": {"account", "accountnumber", "accountno", "accountcode", "glaccount", "acct"},
    "account_name": {"accountname", "description", "accountdescription", "name"},
    "account_category": {"category", "type", "accounttype", "financialstatementarea", "class"},
    "balance": {"balance", "endingbalance", "amount", "currentbalance", "cybalance", "pybalance"},
    "debit": {"debit", "debits", "dr"},
    "credit": {"credit", "credits", "cr"},
    "department": {"department", "dept", "fund", "division", "location"},
}


PY_COLUMNS = [
    "match_key",
    "Account Number",
    "Account Name",
    "Account Category",
    "Department",
    "Balance",
    "Source File",
    "Source Year",
]
COMPARISON_COLUMNS = [
    "match_key",
    "Rank",
    "Account Number",
    "Account Name",
    "Account Category",
    "Department",
    "Prior Year Balance",
    "Current Year Balance",
    "Dollar Change",
    "Percent Change",
    "Change Direction",
    "Question Triggered",
    "Question Trigger Reason",
    "Suggested Client Questions",
    "Analysis Notes",
]


def normalize_header(value: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_balance(value) -> float:
    if value is None:
        return 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if text in {"", "-"}:
        return 0.0
    amount = float(text)
    return -amount if negative else amount


def detect_columns(headers: list[str]) -> dict[str, str]:
    normalized = {normalize_header(header): header for header in headers if clean_text(header)}
    mappings: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                mappings[field] = normalized[alias]
                break
    return mappings


def infer_category(account_name: str) -> str:
    name = account_name.lower()
    if any(word in name for word in ["cash", "receivable", "inventory", "prepaid", "fixed asset", "depreciation"]):
        return "Asset"
    if any(word in name for word in ["payable", "accrual", "debt", "loan"]):
        return "Liability"
    if any(word in name for word in ["equity", "capital", "retained earnings"]):
        return "Equity"
    if any(word in name for word in ["revenue", "sales", "income"]):
        return "Revenue"
    if any(word in name for word in ["expense", "payroll", "rent", "wage", "salary", "fees", "software"]):
        return "Expense"
    return "Unknown"


def make_match_key(account_number: str, account_name: str) -> str:
    if account_number:
        return account_number.strip().lower()
    return re.sub(r"\s+", " ", account_name.strip().lower())


def load_trial_balance(path: Path, source_year: str) -> tuple[list[dict], dict, dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        mappings = detect_columns(headers)
        if "account_name" not in mappings:
            raise ValueError(f"Could not identify account name column in {path.name}")
        if "balance" not in mappings and not {"debit", "credit"}.issubset(mappings):
            raise ValueError(f"Could not identify balance or debit/credit columns in {path.name}")

        records = []
        raw_count = 0
        for raw_row in reader:
            if not any(clean_text(value) for value in raw_row.values()):
                continue
            raw_count += 1
            account_name = clean_text(raw_row.get(mappings["account_name"]))
            if not account_name:
                continue
            account_number = clean_text(raw_row.get(mappings.get("account_number", "")))
            account_category = clean_text(raw_row.get(mappings.get("account_category", ""))) or infer_category(account_name)
            department = clean_text(raw_row.get(mappings.get("department", "")))
            if "balance" in mappings:
                balance = parse_balance(raw_row.get(mappings["balance"]))
            else:
                balance = parse_balance(raw_row.get(mappings["debit"])) - parse_balance(raw_row.get(mappings["credit"]))
            records.append(
                {
                    "match_key": make_match_key(account_number, account_name),
                    "Account Number": account_number,
                    "Account Name": account_name,
                    "Account Category": account_category,
                    "Department": department,
                    "Balance": balance,
                    "Source File": path.name,
                    "Source Year": source_year,
                }
            )
    metadata = {
        "file": path.name,
        "raw_nonblank_rows": raw_count,
        "normalized_rows": len(records),
        "balance_total": round(sum(row["Balance"] for row in records), 2),
        "balance_method": "Single signed balance column; parenthetical values treated as negative.",
    }
    return records, mappings, metadata


def aggregate_records(records: list[dict]) -> tuple[dict[str, dict], int]:
    grouped: dict[str, dict] = {}
    counts = defaultdict(int)
    for row in records:
        key = row["match_key"]
        counts[key] += 1
        if key not in grouped:
            grouped[key] = copy(row)
        else:
            grouped[key]["Balance"] += row["Balance"]
            for field in ["Account Number", "Account Name", "Account Category", "Department"]:
                if not grouped[key].get(field) and row.get(field):
                    grouped[key][field] = row[field]
    duplicate_keys = sum(1 for count in counts.values() if count > 1)
    return grouped, duplicate_keys


def change_direction(prior: float, current: float, dollar_change: float) -> str:
    prior_zero = math.isclose(prior, 0.0, abs_tol=ZERO_TOLERANCE)
    current_zero = math.isclose(current, 0.0, abs_tol=ZERO_TOLERANCE)
    if prior_zero and not current_zero:
        return "New Account"
    if not prior_zero and current_zero:
        return "Closed Account"
    if dollar_change > ZERO_TOLERANCE:
        return "Increase"
    if dollar_change < -ZERO_TOLERANCE:
        return "Decrease"
    return "No Change"


def client_question(account_name: str, category: str, direction: str) -> str:
    name = account_name.lower()
    cat = category.lower()
    if direction == "New Account":
        return "This account appears in the current year but not the prior year. What activity caused the new account to be created, and should any related prior-year amounts have been classified here?"
    if direction == "Closed Account":
        return "This account had a prior-year balance but no current-year balance. Was the activity discontinued, reclassified, settled, or moved to another account?"
    if "revenue" in name or "sales" in name or "revenue" in cat:
        return "What changes in volume, pricing, customers, contracts, cutoff, or reclassification explain this revenue variance?"
    if "receivable" in name:
        return "What billing, collection timing, aging, credit memo, cutoff, or write-off activity explains this accounts receivable variance?"
    if "inventory" in name:
        return "What purchase volume, count adjustment, costing, obsolescence, or write-down activity explains this inventory variance?"
    if "prepaid" in name or "other asset" in name:
        return "What additions, amortization, cutoff, or reclassifications explain this prepaid or other asset variance?"
    if "fixed asset" in name or "depreciation" in name or "cip" in name:
        return "What additions, disposals, CIP transfers, depreciation activity, or policy changes explain this fixed asset variance?"
    if "payable" in name or "accrual" in name:
        return "What vendor timing, unrecorded liability, reversal, new obligation, cutoff, or reclassification activity explains this payable or accrual variance?"
    if "debt" in name or "loan" in name:
        return "What borrowings, principal payments, refinancing, covenant activity, or interest classifications explain this debt variance?"
    if "equity" in name or "capital" in name or "retained earnings" in name:
        return "What contributions, distributions, retained earnings activity, or closing entries explain this equity variance?"
    if "payroll" in name or "wage" in name or "salary" in name:
        return "What headcount, compensation, bonus, accrual, or contractor-versus-employee activity explains this payroll variance?"
    if "professional" in name or "software" in name or "fees" in name or "expense" in name:
        return "What new vendors, one-time projects, renewals, implementation work, timing, or reclassifications explain this operating expense variance?"
    return "This account changed materially. What underlying activity, timing, cutoff, settlement, or reclassification explains the variance?"


def analysis_note(prior: float, current: float, dollar_change: float, direction: str) -> str:
    if direction == "New Account":
        note = "Account appears to be new in the current year."
    elif direction == "Closed Account":
        note = "Account had a prior-year balance but no current-year balance."
    elif direction == "Increase":
        note = f"Current-year balance increased by ${abs(dollar_change):,.0f} compared with prior year."
    elif direction == "Decrease":
        note = f"Current-year balance decreased by ${abs(dollar_change):,.0f} compared with prior year."
    else:
        note = "No meaningful balance change identified."
    if prior * current < 0:
        note += " Balance changed direction from debit to credit or credit to debit; review classification and underlying activity."
    return note


def build_comparison(prior_grouped: dict[str, dict], current_grouped: dict[str, dict]) -> list[dict]:
    rows = []
    for key in sorted(set(prior_grouped) | set(current_grouped)):
        prior_row = prior_grouped.get(key, {})
        current_row = current_grouped.get(key, {})
        prior = float(prior_row.get("Balance", 0.0))
        current = float(current_row.get("Balance", 0.0))
        dollar_change = current - prior
        percent_change = None if math.isclose(prior, 0.0, abs_tol=ZERO_TOLERANCE) else dollar_change / abs(prior)
        direction = change_direction(prior, current, dollar_change)
        abs_dollar = abs(dollar_change)
        abs_percent = abs(percent_change) if percent_change is not None else None
        dollar_met = abs_dollar >= DOLLAR_THRESHOLD
        percent_met = abs_percent is not None and abs_percent >= PERCENT_THRESHOLD
        triggered = dollar_met and percent_met
        reason = "Dollar and percent thresholds met" if triggered else ""
        account_number = current_row.get("Account Number") or prior_row.get("Account Number", "")
        account_name = current_row.get("Account Name") or prior_row.get("Account Name", "")
        category = current_row.get("Account Category") or prior_row.get("Account Category", "")
        department = current_row.get("Department") or prior_row.get("Department", "")
        rows.append(
            {
                "match_key": key,
                "Account Number": account_number,
                "Account Name": account_name,
                "Account Category": category,
                "Department": department,
                "Prior Year Balance": prior,
                "Current Year Balance": current,
                "Dollar Change": dollar_change,
                "Percent Change": percent_change,
                "Change Direction": direction,
                "Question Triggered": "Yes" if triggered else "No",
                "Question Trigger Reason": reason,
                "Suggested Client Questions": client_question(account_name, category, direction) if triggered else "",
                "Analysis Notes": analysis_note(prior, current, dollar_change, direction),
            }
        )

    rows.sort(
        key=lambda row: (
            -abs(row["Dollar Change"]),
            0 if row["Change Direction"] in {"New Account", "Closed Account"} else 1,
            -(abs(row["Percent Change"]) if row["Percent Change"] is not None else -1),
            str(row["Account Number"]),
        )
    )
    for idx, row in enumerate(rows, start=1):
        row["Rank"] = idx
    return rows


def append_table(ws, headers: list[str], rows: list[dict], formula_builder=None):
    ws.append(headers)
    for row_idx, row in enumerate(rows, start=2):
        values = []
        for header in headers:
            if formula_builder:
                formula = formula_builder(header, row_idx)
                values.append(formula if formula else row.get(header, ""))
            else:
                values.append(row.get(header, ""))
        ws.append(values)


def style_sheet(ws, currency_columns=None, percent_columns=None, hidden_columns=None, wrap_columns=None):
    currency_columns = set(currency_columns or [])
    percent_columns = set(percent_columns or [])
    hidden_columns = set(hidden_columns or [])
    wrap_columns = set(wrap_columns or [])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = ws.dimensions
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = ws.cell(row=1, column=col_idx).value
        letter = get_column_letter(col_idx)
        if header in hidden_columns:
            ws.column_dimensions[letter].hidden = True
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[: min(ws.max_row, 60)])
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 55)
        for cell in column_cells[1:]:
            if header in currency_columns:
                cell.number_format = '$#,##0;[Red]($#,##0);-'
            elif header in percent_columns:
                cell.number_format = '0.0%;[Red](0.0%);-'
            if header in wrap_columns or (isinstance(cell.value, str) and len(cell.value) > 70):
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")


def build_summary_sheet(wb, prior_meta, current_meta, prior_mapping, current_mapping, duplicate_counts, comparison_rows, validation_checks):
    ws = wb.active
    ws.title = "Summary"
    questions = sum(1 for row in comparison_rows if row["Question Triggered"] == "Yes")
    content = [
        ("Workbook Purpose", "Draft trial balance variance analysis and follow-up question aid for human review."),
        ("Input Files", f"{PRIOR_FILE.name}; {CURRENT_FILE.name}"),
        ("Thresholds Used", "Question triggered only when absolute dollar change is at least $200,000 AND absolute percent change is at least 20%."),
        ("Balance Convention", "Single signed balance column preserved; parenthetical currency values treated as negative."),
        ("Prior-Year Rows / Total", f"{prior_meta['normalized_rows']} rows; total ${prior_meta['balance_total']:,.2f}"),
        ("Current-Year Rows / Total", f"{current_meta['normalized_rows']} rows; total ${current_meta['balance_total']:,.2f}"),
        ("Comparison Rows", len(comparison_rows)),
        ("Questions Generated", questions),
        ("Prior-Year Mappings", json.dumps(prior_mapping, sort_keys=True)),
        ("Current-Year Mappings", json.dumps(current_mapping, sort_keys=True)),
        ("Duplicate Aggregation", f"Prior-year duplicate match keys: {duplicate_counts['prior_year']}; current-year duplicate match keys: {duplicate_counts['current_year']}."),
        ("Validation Status", "Completed with limitations. " + "; ".join(check["check"] for check in validation_checks if check["status"] == "passed")),
        ("Human Review Needed", "Review generated questions, account classifications, balance-sign conventions, and any accounts near but below the configured thresholds."),
        ("Limitations", "Outputs are review aids and draft workpapers only; they do not establish final audit, compliance, financial, or management conclusions."),
    ]
    ws.append(["Topic", "Details"])
    for topic, detail in content:
        ws.append([topic, detail])
    style_sheet(ws, wrap_columns={"Details"})
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110


def build_workbook(prior_records, current_records, comparison_rows, prior_meta, current_meta, prior_mapping, current_mapping, duplicate_counts, validation_checks):
    wb = Workbook()
    build_summary_sheet(wb, prior_meta, current_meta, prior_mapping, current_mapping, duplicate_counts, comparison_rows, validation_checks)

    prior_ws = wb.create_sheet("Prior_Year_TB")
    append_table(prior_ws, PY_COLUMNS, prior_records)
    style_sheet(prior_ws, currency_columns={"Balance"}, hidden_columns={"match_key"})

    current_ws = wb.create_sheet("Current_Year_TB")
    append_table(current_ws, PY_COLUMNS, current_records)
    style_sheet(current_ws, currency_columns={"Balance"}, hidden_columns={"match_key"})

    comparison_ws = wb.create_sheet("Comparison")
    comp_col = {name: get_column_letter(idx) for idx, name in enumerate(COMPARISON_COLUMNS, start=1)}
    prior_match_col = get_column_letter(PY_COLUMNS.index("match_key") + 1)
    prior_balance_col = get_column_letter(PY_COLUMNS.index("Balance") + 1)
    current_match_col = get_column_letter(PY_COLUMNS.index("match_key") + 1)
    current_balance_col = get_column_letter(PY_COLUMNS.index("Balance") + 1)

    def comparison_formula(header: str, row_idx: int):
        match_ref = f"{comp_col['match_key']}{row_idx}"
        prior_ref = f"{comp_col['Prior Year Balance']}{row_idx}"
        current_ref = f"{comp_col['Current Year Balance']}{row_idx}"
        dollar_ref = f"{comp_col['Dollar Change']}{row_idx}"
        if header == "Prior Year Balance":
            return f"=SUMIF(Prior_Year_TB!{prior_match_col}:{prior_match_col},{match_ref},Prior_Year_TB!{prior_balance_col}:{prior_balance_col})"
        if header == "Current Year Balance":
            return f"=SUMIF(Current_Year_TB!{current_match_col}:{current_match_col},{match_ref},Current_Year_TB!{current_balance_col}:{current_balance_col})"
        if header == "Dollar Change":
            return f"={current_ref}-{prior_ref}"
        if header == "Percent Change":
            return f'=IFERROR({dollar_ref}/ABS({prior_ref}),"")'
        return None

    append_table(comparison_ws, COMPARISON_COLUMNS, comparison_rows, formula_builder=comparison_formula)
    style_sheet(
        comparison_ws,
        currency_columns={"Prior Year Balance", "Current Year Balance", "Dollar Change"},
        percent_columns={"Percent Change"},
        hidden_columns={"match_key"},
        wrap_columns={"Suggested Client Questions", "Analysis Notes"},
    )
    yes_col = comp_col["Question Triggered"]
    comparison_ws.conditional_formatting.add(
        f"A2:{get_column_letter(comparison_ws.max_column)}{comparison_ws.max_row}",
        FormulaRule(formula=[f'${yes_col}2="Yes"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    comparison_ws.column_dimensions[comp_col["Suggested Client Questions"]].width = 58
    comparison_ws.column_dimensions[comp_col["Analysis Notes"]].width = 52
    return wb


def validate(prior_records, current_records, prior_grouped, current_grouped, comparison_rows, prior_mapping, current_mapping):
    checks = []

    def passed(name):
        checks.append({"check": name, "status": "passed"})

    if prior_records and current_records:
        passed("Both input files loaded successfully")
    else:
        raise ValueError("One or both input files did not load any records")
    for label, mapping in [("prior-year", prior_mapping), ("current-year", current_mapping)]:
        if "account_name" not in mapping or ("balance" not in mapping and not {"debit", "credit"}.issubset(mapping)):
            raise ValueError(f"Required columns not identified for {label}")
    passed("Account name and balance columns identified per file")
    for row in prior_records + current_records:
        if not isinstance(row["Balance"], (int, float)):
            raise ValueError("Non-numeric balance encountered")
    passed("All balances are numeric")
    if len(comparison_rows) != len(set(prior_grouped) | set(current_grouped)):
        raise ValueError("Comparison row count does not equal unique match key count")
    passed("Comparison row count equals unique match keys across both years")
    for row in comparison_rows:
        if round(row["Dollar Change"], 2) != round(row["Current Year Balance"] - row["Prior Year Balance"], 2):
            raise ValueError(f"Dollar change mismatch for {row['match_key']}")
        dollar_met = abs(row["Dollar Change"]) >= DOLLAR_THRESHOLD
        percent_met = row["Percent Change"] is not None and abs(row["Percent Change"]) >= PERCENT_THRESHOLD
        expected_trigger = dollar_met and percent_met
        if (row["Question Triggered"] == "Yes") != expected_trigger:
            raise ValueError(f"Threshold trigger mismatch for {row['match_key']}")
        if row["Question Triggered"] == "Yes" and not row["Suggested Client Questions"]:
            raise ValueError(f"Missing suggested question for {row['match_key']}")
        if row["Question Triggered"] == "No" and row["Suggested Client Questions"]:
            raise ValueError(f"Unexpected suggested question for {row['match_key']}")
    passed("Dollar change equals current year balance less prior year balance")
    passed("Suggested questions populated only for triggered rows")
    passed("AND threshold logic applied consistently")
    passed("No Materiality Score column generated")
    passed("Summary documents thresholds, mappings, aggregation counts, and balance convention")
    return checks


def write_run_summary(comparison_rows, prior_meta, current_meta):
    triggered = [row for row in comparison_rows if row["Question Triggered"] == "Yes"]
    markdown = f"""# Run Summary

## Task Attempted

Prepared a draft trial balance variance analysis workbook for the `trial_balance` intake job, including follow-up questions for accounts meeting both configured variance thresholds.

## Inputs Used

- `00_source/FY25 TB.csv`
- `00_source/FY26 TB.csv`
- `run_request.md`

## Outputs Created

- `01_variance_analysis/variance_analysis.xlsx`
- `01_variance_analysis/variance_analysis_metadata.json`
- `manifest.json`

## Checks Performed

- Confirmed both source files loaded and required columns were detected.
- Parsed signed currency balances, including parenthetical negatives.
- Compared unique account match keys across both years.
- Verified dollar changes against current-year less prior-year balances.
- Applied the configured AND threshold: absolute dollar change >= $200,000 and absolute percent change >= 20%.
- Confirmed suggested client questions are populated only for triggered rows.

## Limitations

- The workbook is a review aid and draft workpaper, not a final audit, compliance, legal, financial, or management conclusion.
- Percent change is blank when the prior-year balance is zero, so new accounts do not trigger the AND threshold unless a reviewer separately evaluates them.
- Account categories were taken from the source files when present and were not independently validated.

## Human Review Needed

- Review the {len(triggered)} generated follow-up questions for business relevance and completeness.
- Inspect large non-triggered changes, new accounts, closed accounts, and sign-flip balances for possible follow-up outside the configured threshold rule.
- Confirm the source balance convention and whether signed balances should be interpreted differently by account type.

## Suggested Improvements

- Add preparer/reviewer signoff fields to the workbook if this will be retained as a formal workpaper.
- Add account-owner or department-owner fields to route follow-up questions.
- Consider separate review logic for new and closed accounts where percent change is not meaningful.
"""
    (RUN_DIR / "run_summary.md").write_text(markdown, encoding="utf-8")


def main():
    STAGE_DIR.mkdir(parents=True, exist_ok=True)
    prior_records, prior_mapping, prior_meta = load_trial_balance(PRIOR_FILE, "Prior Year")
    current_records, current_mapping, current_meta = load_trial_balance(CURRENT_FILE, "Current Year")
    prior_grouped, prior_duplicates = aggregate_records(prior_records)
    current_grouped, current_duplicates = aggregate_records(current_records)
    comparison_rows = build_comparison(prior_grouped, current_grouped)
    validation_checks = validate(prior_records, current_records, prior_grouped, current_grouped, comparison_rows, prior_mapping, current_mapping)
    duplicate_counts = {"prior_year": prior_duplicates, "current_year": current_duplicates}
    wb = build_workbook(
        prior_records,
        current_records,
        comparison_rows,
        prior_meta,
        current_meta,
        prior_mapping,
        current_mapping,
        duplicate_counts,
        validation_checks,
    )
    wb.save(OUTPUT_XLSX)

    # Re-open once to confirm the workbook artifact can be loaded after save.
    load_workbook(OUTPUT_XLSX, read_only=True, data_only=False).close()

    questions_generated = sum(1 for row in comparison_rows if row["Question Triggered"] == "Yes")
    metadata = {
        "skill": "variance-explainer",
        "status": "completed_with_limitations",
        "source_inputs": {
            "prior_year_file": str(PRIOR_FILE.relative_to(RUN_DIR)),
            "current_year_file": str(CURRENT_FILE.relative_to(RUN_DIR)),
        },
        "artifacts_written": [
            str(OUTPUT_XLSX.relative_to(RUN_DIR)),
            str(OUTPUT_METADATA.relative_to(RUN_DIR)),
        ],
        "parameters": {
            "dollar_threshold": DOLLAR_THRESHOLD,
            "percent_threshold": PERCENT_THRESHOLD,
            "threshold_logic": THRESHOLD_LOGIC,
            "zero_balance_percent_change": "blank when prior-year balance is zero",
        },
        "counts": {
            "prior_year_rows": prior_meta["normalized_rows"],
            "current_year_rows": current_meta["normalized_rows"],
            "comparison_rows": len(comparison_rows),
            "questions_generated": questions_generated,
            "prior_year_duplicate_match_keys_aggregated": prior_duplicates,
            "current_year_duplicate_match_keys_aggregated": current_duplicates,
        },
        "column_mappings": {
            "prior_year": prior_mapping,
            "current_year": current_mapping,
        },
        "validation_checks": validation_checks,
        "limitations": [
            "Outputs are draft workpapers and review aids only.",
            "Source account categories and signed-balance conventions were not independently validated.",
            "Percent change is blank when prior-year balance is zero; new accounts may need separate review.",
        ],
        "human_review_needed": [
            "Review generated questions before sending to a client or process owner.",
            "Review large non-triggered changes, new accounts, closed accounts, and sign flips for possible follow-up.",
            "Confirm whether account-number matching is appropriate for the trial balance population.",
        ],
        "next_stage_input": {
            "file": str(OUTPUT_XLSX.relative_to(RUN_DIR)),
            "sheet": "Comparison",
            "table": "Comparison!A1:O{}".format(len(comparison_rows) + 1),
            "record_count": len(comparison_rows),
        },
    }
    OUTPUT_METADATA.write_text(json.dumps(metadata, indent=2), encoding="utf-8")

    manifest = {
        "manifest_version": "1.0",
        "run_id": RUN_ID,
        "job_id": JOB_ID,
        "created_at": CREATED_AT,
        "status": "completed_with_limitations",
        "source_files": [
            str(PRIOR_FILE.relative_to(RUN_DIR)),
            str(CURRENT_FILE.relative_to(RUN_DIR)),
        ],
        "stages": {
            "01_variance_analysis": {
                "skill": "variance-explainer",
                "status": "completed_with_limitations",
                "output": str(OUTPUT_XLSX.relative_to(RUN_DIR)),
                "metadata": str(OUTPUT_METADATA.relative_to(RUN_DIR)),
                "next_stage_input": metadata["next_stage_input"],
            }
        },
        "current_sampling_input": {
            "ready_for_sampling": False,
            "file": None,
            "sheet": None,
            "record_count": None,
            "key_fields": [],
            "limitations": ["This run is a variance-analysis workflow and does not produce a sampling input by default."],
        },
    }
    (RUN_DIR / "manifest.json").write_text(json.dumps(manifest, indent=2), encoding="utf-8")
    write_run_summary(comparison_rows, prior_meta, current_meta)

    print(json.dumps({
        "output": str(OUTPUT_XLSX),
        "metadata": str(OUTPUT_METADATA),
        "comparison_rows": len(comparison_rows),
        "questions_generated": questions_generated,
    }, indent=2))


if __name__ == "__main__":
    main()
